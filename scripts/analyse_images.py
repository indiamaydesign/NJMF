#!/usr/bin/env python3
"""
Use Claude vision to read each archival photo, detect box/folder separators,
and copy documents into sorted folders.

Convention: before photographing documents, a photo is taken of the physical
box or cartridge label. This script detects those separator photos and copies
the subsequent document photos into public/images/primary/sorted/<box-name>/.

Saves progress to data/analysis_progress.json so it can be safely restarted.
Final output also updates data/image_records.csv and data/image_records.xlsx.

Usage:
    cd scripts
    python3 analyse_images.py
    python3 analyse_images.py --dry-run   # analyse without copying files

Set ANTHROPIC_API_KEY in your environment before running.
"""

import argparse
import base64
import csv
import io
import json
import os
import shutil
import time
from pathlib import Path

import anthropic
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image
import pillow_heif

pillow_heif.register_heif_opener()

ROOT = Path(__file__).parent.parent
IMAGES_DIR = ROOT / "public" / "images"
DATA_DIR = ROOT / "data"
CSV_PATH = DATA_DIR / "image_records.csv"
XLSX_PATH = DATA_DIR / "image_records.xlsx"
PROGRESS_PATH = DATA_DIR / "analysis_progress.json"
SORTED_DIR = ROOT / "public" / "images" / "primary" / "sorted"

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".heic", ".webp"}
MAX_SIDE_PX = 1024

COLUMNS = [
    "filename", "source_device", "capture_date", "year_guess",
    "gps_lat", "gps_lon", "camera_model",
    "archival_box", "is_separator", "document_date", "description",
    "file_path", "notes",
]

PROMPT = """You are analysing a photograph of an archival document from a Norwegian research project (1969–1975).

Your task is to find the DATE THIS DOCUMENT WAS WRITTEN OR ISSUED — not dates mentioned inside the text.

Rules for finding the document date:
- Look at the TOP of the document first: the date is usually in the header, subject line, or top-right corner
- For letters (brev): date appears near the top, often right-aligned (e.g. "Oslo, 15. mai 1972")
- For meeting minutes (møtereferat, protokoll, referat): look for the meeting date in the header block
- For reports: look for the date on the cover page or first line
- Norwegian date format: day.month.year — e.g. "15.06.1974" means 15 June 1974, NOT 6 March 1974
- IGNORE dates mentioned in the body text as historical references, statistics, or ranges (e.g. "since 1965", "founded in 1938", "the period 1970–1975")
- If you see multiple dates, choose only the one that is clearly the document's own issue date
- If no clear document date is visible at the top, return ""

Look carefully at the image and respond in JSON with exactly these fields:

{
  "is_separator": true/false,   // true if this photo shows a physical box, folder, or cartridge label/cover (not a document itself)
  "box_name": "...",            // if is_separator is true: the box or folder name/number as written on the label. Otherwise ""
  "document_date": "...",       // the date this document was issued, from the top of the page. Single date only, in the format shown. "" if not visible
  "description": "..."          // one sentence: document type, subject, any people or organisations named
}

Respond with valid JSON only. No prose outside the JSON.
"""


def heic_to_jpeg_b64(path: Path) -> str:
    img = Image.open(path)
    img = img.convert("RGB")
    w, h = img.size
    if max(w, h) > MAX_SIDE_PX:
        ratio = MAX_SIDE_PX / max(w, h)
        img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return base64.standard_b64encode(buf.getvalue()).decode("utf-8")


def analyse_image(client: anthropic.Anthropic, path: Path) -> dict:
    b64 = heic_to_jpeg_b64(path)
    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=256,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
                    {"type": "text", "text": PROMPT},
                ],
            }
        ],
    )
    raw = message.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


def copy_to_sorted(path: Path, box_name: str, dry_run: bool) -> Path:
    folder = box_name if box_name else "_unsorted"
    dest_dir = SORTED_DIR / folder
    dest = dest_dir / path.name
    if not dry_run:
        dest_dir.mkdir(parents=True, exist_ok=True)
        if not dest.exists():
            shutil.copy2(path, dest)
    return dest


def load_csv() -> list[dict]:
    with open(CSV_PATH, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_progress() -> dict:
    if PROGRESS_PATH.exists():
        with open(PROGRESS_PATH) as f:
            return json.load(f)
    return {}


def save_progress(progress: dict) -> None:
    with open(PROGRESS_PATH, "w") as f:
        json.dump(progress, f, indent=2)


def write_csv(records: list[dict]) -> None:
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)


def write_xlsx(records: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Image Records"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(records, start=2):
        for col_idx, col in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(col, ""))

    widths = {
        "filename": 28, "source_device": 12, "capture_date": 24, "year_guess": 10,
        "gps_lat": 12, "gps_lon": 12, "camera_model": 16,
        "archival_box": 24, "is_separator": 13, "document_date": 18,
        "description": 50, "file_path": 52, "notes": 30,
    }
    for col_idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = widths.get(col, 15)

    ws.freeze_panes = "A2"
    wb.save(XLSX_PATH)


def recover_current_box(records: list[dict], progress: dict) -> str:
    """Find the box name from the last separator already in progress."""
    current_box = ""
    for record in records:
        if record["filename"] not in progress:
            break
        result = progress[record["filename"]]
        if result.get("is_separator"):
            current_box = result.get("box_name", "")
    return current_box


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyse and sort archival photos using Claude vision.")
    parser.add_argument("--dry-run", action="store_true", help="Analyse without copying files")
    args = parser.parse_args()

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("Error: ANTHROPIC_API_KEY environment variable not set.")
        print("Run: export ANTHROPIC_API_KEY=your_key_here")
        return

    client = anthropic.Anthropic(api_key=api_key)
    records = load_csv()
    progress = load_progress()

    for r in records:
        r.setdefault("document_date", "")

    todo = [r for r in records if r["filename"] not in progress]
    done_count = len(records) - len(todo)
    print(f"{len(records)} images total. {done_count} already done, {len(todo)} to analyse.")
    if args.dry_run:
        print("DRY RUN — no files will be copied.")

    current_box = recover_current_box(records, progress)
    if current_box:
        print(f"Resuming in box: {current_box}")

    for i, record in enumerate(todo, start=1):
        path = ROOT / record["file_path"]
        print(f"  [{i}/{len(todo)}] {record['filename']}...", end=" ", flush=True)

        try:
            result = analyse_image(client, path)
        except Exception as e:
            print(f"ERROR: {e}")
            progress[record["filename"]] = {"error": str(e)}
            save_progress(progress)
            continue

        if result.get("is_separator"):
            current_box = result.get("box_name", "")
            record["is_separator"] = "Y"
            record["archival_box"] = current_box
            dest = copy_to_sorted(path, current_box, args.dry_run)
            print(f"SEPARATOR -> {current_box or '(unnamed)'} [{dest}]")
        else:
            record["is_separator"] = ""
            record["archival_box"] = current_box
            record["document_date"] = result.get("document_date", "")
            record["description"] = result.get("description", "")
            dest = copy_to_sorted(path, current_box, args.dry_run)
            print(f"{result.get('document_date', '?')} | {result.get('description', '')[:60]}")

        progress[record["filename"]] = result
        save_progress(progress)

        if i % 25 == 0:
            write_csv(records)
            write_xlsx(records)
            print(f"  -- checkpoint saved ({done_count + i} total done) --")

        time.sleep(0.1)

    write_csv(records)
    write_xlsx(records)
    print(f"\nDone. Sorted files in {SORTED_DIR}")
    print(f"Records in {XLSX_PATH}")


if __name__ == "__main__":
    main()
