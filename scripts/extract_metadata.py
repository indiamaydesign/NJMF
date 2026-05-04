#!/usr/bin/env python3
"""
Extract metadata from all archival images and produce a tracking spreadsheet.

Uses macOS `mdls` to read HEIC/image metadata (capture date, GPS, camera model).

Output: data/image_records.csv and data/image_records.xlsx

Columns:
    filename         - original filename
    source_device    - iphone / android / unknown
    capture_date     - photo capture datetime
    year_guess       - year extracted from capture_date
    gps_lat          - GPS latitude if available
    gps_lon          - GPS longitude if available
    camera_model     - camera model from EXIF
    archival_box     - populated after sort_images.py is run (blank for now)
    is_separator     - mark Y for folder-cover photos
    description      - blank, fill in as you work through the archive
    file_path        - relative path from NJMF root
    notes            - any extra notes

Usage:
    cd scripts
    python3 extract_metadata.py
"""

import csv
import re
import subprocess
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).parent.parent
IMAGES_DIR = ROOT / "public" / "images"
OUTPUT_DIR = ROOT / "data"
CSV_OUT = OUTPUT_DIR / "image_records.csv"
XLSX_OUT = OUTPUT_DIR / "image_records.xlsx"

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".heic", ".webp"}

MDLS_KEYS = [
    "kMDItemContentCreationDate",
    "kMDItemGPSLatitude",
    "kMDItemGPSLongitude",
    "kMDItemAcquisitionModel",
]

COLUMNS = [
    "filename",
    "source_device",
    "capture_date",
    "year_guess",
    "gps_lat",
    "gps_lon",
    "camera_model",
    "archival_box",
    "is_separator",
    "description",
    "file_path",
    "notes",
]


def mdls(path: Path) -> dict[str, str]:
    args = ["mdls"] + sum([["-name", k] for k in MDLS_KEYS], []) + [str(path)]
    result = subprocess.run(args, capture_output=True, text=True)
    out: dict[str, str] = {}
    for line in result.stdout.splitlines():
        if "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip('"')
        if value and value != "(null)":
            out[key] = value
    return out


def infer_device(path: Path) -> str:
    parts = path.parts
    if "iphone" in parts:
        return "iphone"
    if "android" in parts:
        return "android"
    return "unknown"


def extract_year(date_str: str) -> str:
    match = re.search(r"\b(19|20)\d{2}\b", date_str)
    return match.group(0) if match else ""


def scan_images() -> list[dict]:
    paths = [
        p for p in IMAGES_DIR.rglob("*")
        if p.suffix.lower() in IMAGE_EXTENSIONS
    ]
    print(f"Found {len(paths)} images. Extracting metadata...")

    records = []
    for i, path in enumerate(paths, 1):
        if i % 50 == 0:
            print(f"  {i}/{len(paths)}...")
        meta = mdls(path)
        capture_date = meta.get("kMDItemContentCreationDate", "")
        records.append({
            "filename": path.name,
            "source_device": infer_device(path),
            "capture_date": capture_date,
            "year_guess": extract_year(capture_date),
            "gps_lat": meta.get("kMDItemGPSLatitude", ""),
            "gps_lon": meta.get("kMDItemGPSLongitude", ""),
            "camera_model": meta.get("kMDItemAcquisitionModel", ""),
            "archival_box": "",
            "is_separator": "",
            "description": "",
            "file_path": str(path.relative_to(ROOT)),
            "notes": "",
        })

    records.sort(key=lambda r: (r["capture_date"] == "", r["capture_date"], r["filename"]))
    return records


def write_csv(records: list[dict]) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(records)
    print(f"CSV written: {CSV_OUT}")


def write_xlsx(records: list[dict]) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Image Records"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(records, start=2):
        for col_idx, col in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=record[col])

    widths = {
        "filename": 28, "source_device": 14, "capture_date": 24, "year_guess": 10,
        "gps_lat": 14, "gps_lon": 14, "camera_model": 18,
        "archival_box": 22, "is_separator": 13,
        "description": 44, "file_path": 52, "notes": 30,
    }
    for col_idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = widths.get(col, 15)

    ws.freeze_panes = "A2"
    wb.save(XLSX_OUT)
    print(f"Excel written: {XLSX_OUT}")


def main() -> None:
    records = scan_images()
    dated = sum(1 for r in records if r["capture_date"])
    print(f"\n{dated}/{len(records)} images have a capture date.")
    write_csv(records)
    write_xlsx(records)
    print("\nDone. Open data/image_records.xlsx to start filling in archival_box and descriptions.")


if __name__ == "__main__":
    main()
